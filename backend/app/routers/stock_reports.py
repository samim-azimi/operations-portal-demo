import csv
import io
from calendar import month_name
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    AuditLog, OrganizationSettings, PhysicalInventoryCount, StockCard,
    StockItem, StockMovement, User,
)
from app.security import require_permission
from app.services.pdf_reports import build_stock_card_pdf

router = APIRouter(
    prefix="/stock/reports",
    tags=["Stock Reports"],
    dependencies=[Depends(require_permission("can_access_stock"))],
)


def get_card(card_id:int,db:Session):
    card=db.get(StockCard,card_id)
    if not card or not card.is_active:raise HTTPException(404,"Stock card not found")
    return card


def stock_card_report(card_id:int,date_from:date,date_to:date,db:Session):
    if date_to<date_from:raise HTTPException(422,"To Date cannot be earlier than From Date")
    card=get_card(card_id,db)
    previous=float(card.opening_quantity or 0)+float(
        db.query(func.coalesce(func.sum(StockMovement.quantity_in-StockMovement.quantity_out),0))
        .filter(StockMovement.stock_card_id==card.id,StockMovement.movement_date<date_from).scalar() or 0
    )
    movements=db.query(StockMovement).filter(
        StockMovement.stock_card_id==card.id,
        StockMovement.movement_date>=date_from,
        StockMovement.movement_date<=date_to,
    ).order_by(StockMovement.movement_date,StockMovement.id).all()
    balance=previous;rows=[]
    for movement in movements:
        balance+=float(movement.quantity_in or 0)-float(movement.quantity_out or 0)
        rows.append({
            "date":movement.movement_date.strftime("%d/%m/%Y"),
            "goods_received_note_no":movement.goods_received_note_no,
            "in":movement.quantity_in,"out":movement.quantity_out,
            "stock_transfer_no":movement.stock_transfer_no,
            "balance":balance,"destination":movement.destination,
            "remarks":movement.remarks or movement.comments,
            "signature":movement.signature_name,
        })
    return card,previous,movements,rows,balance


@router.get("/stock-card/preview")
def preview_stock_card(
    card_id:int,from_date:date,to_date:date,db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_stock_card")),
):
    card,previous,_,rows,final=stock_card_report(card_id,from_date,to_date,db)
    db.add(AuditLog(actor_id=actor.id,action="Stock Card preview generated",details={"stock_card_id":card_id,"from":str(from_date),"to":str(to_date)}));db.commit()
    return {"stock_card":{"id":card.id,"number":card.stock_card_number,"item":card.item_name,"base":card.base,"unit":card.unit,"donor":card.donor,"project_code":card.project_code},"previous_balance":previous,"transactions":rows,"final_balance":final,"message":None if rows else "No transactions found for this period"}


@router.get("/stock-card/export/pdf")
def export_stock_card_pdf(
    card_id:int,from_date:date,to_date:date,db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_stock_card")),
):
    card,previous,movements,_,_=stock_card_report(card_id,from_date,to_date,db)
    organization=db.query(OrganizationSettings).first()
    content=build_stock_card_pdf(card,movements,previous,from_date,to_date,organization)
    db.add(AuditLog(actor_id=actor.id,action="Stock Card PDF exported",details={"stock_card_id":card_id,"from":str(from_date),"to":str(to_date)}));db.commit()
    filename=f"StockCard_{card.item_name.replace(' ','_')}_{from_date}_to_{to_date}.pdf"
    return StreamingResponse(io.BytesIO(content),media_type="application/pdf",headers={"Content-Disposition":f'attachment; filename="{filename}"',"X-Content-Type-Options":"nosniff"})


def movement_query(db,mission=None,base=None,year=None,from_date=None,to_date=None,stock_card_number=None,donor=None,project_code=None,item=None,month_number=None):
    query=db.query(StockMovement,StockCard,StockItem).join(StockItem,StockMovement.stock_item_id==StockItem.id).outerjoin(StockCard,StockMovement.stock_card_id==StockCard.id)
    if mission:query=query.filter(StockMovement.mission==mission)
    if base:query=query.filter(StockMovement.base==base)
    if year:query=query.filter(StockMovement.year==year)
    if from_date:query=query.filter(StockMovement.movement_date>=from_date)
    if to_date:query=query.filter(StockMovement.movement_date<=to_date)
    if stock_card_number:query=query.filter(StockCard.stock_card_number==stock_card_number)
    if donor:query=query.filter(StockCard.donor==donor)
    if project_code:query=query.filter(StockCard.project_code==project_code)
    if item:query=query.filter(StockItem.item_name.ilike(f"%{item}%"))
    if month_number:query=query.filter(StockMovement.month_number==month_number)
    return query


def movement_row(movement,card,item):
    return {
        "Stock Card Number":card.stock_card_number if card else "",
        "Donor":card.donor if card else item.donor or "",
        "Project Code":card.project_code if card else item.project_code or "",
        "Item":item.item_name,"Specifications":item.specifications or "","Unit":item.unit,
        "Date mouvement JJ/MM/AAAA":movement.movement_date.strftime("%d/%m/%Y"),
        "Month Number":movement.month_number,"IN":movement.quantity_in or 0,"OUT":movement.quantity_out or 0,
        "PO Number":movement.po_number or "","Waybill Number":movement.waybill_number or "",
        "Comments":movement.comments or movement.remarks or "",
    }


@router.get("/movements/preview")
def preview_movements(
    mission:str|None=None,base:str|None=None,year:int|None=None,from_date:date|None=None,to_date:date|None=None,
    stock_card_number:str|None=None,donor:str|None=None,project_code:str|None=None,item:str|None=None,
    month_number:int|None=Query(None,ge=1,le=12),page:int=Query(1,ge=1),page_size:int=Query(50,ge=1,le=200),
    db:Session=Depends(get_db),actor:User=Depends(require_permission("can_export_stock_movements")),
):
    query=movement_query(db,mission,base,year,from_date,to_date,stock_card_number,donor,project_code,item,month_number)
    total=query.count();records=query.order_by(StockMovement.movement_date,StockMovement.id).offset((page-1)*page_size).limit(page_size).all()
    db.add(AuditLog(actor_id=actor.id,action="STOCK FU : MOUVEMENTS previewed",details={"year":year,"records":total}));db.commit()
    return {"items":[movement_row(*record) for record in records],"total":total,"page":page,"page_size":page_size,"pages":(total+page_size-1)//page_size}


@router.get("/movements/export/csv")
def export_movements(
    mission:str|None=None,base:str|None=None,year:int|None=None,from_date:date|None=None,to_date:date|None=None,
    stock_card_number:str|None=None,donor:str|None=None,project_code:str|None=None,item:str|None=None,
    month_number:int|None=Query(None,ge=1,le=12),db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_stock_movements")),
):
    records=movement_query(db,mission,base,year,from_date,to_date,stock_card_number,donor,project_code,item,month_number).order_by(StockMovement.movement_date,StockMovement.id).all()
    headers=["Stock Card Number","Donor","Project Code","Item","Specifications","Unit","Date mouvement JJ/MM/AAAA","Month Number","IN","OUT","PO Number","Waybill Number","Comments"]
    output=io.StringIO();writer=csv.DictWriter(output,fieldnames=headers);writer.writeheader()
    for record in records:writer.writerow(movement_row(*record))
    db.add(AuditLog(actor_id=actor.id,action="STOCK FU : MOUVEMENTS CSV exported",details={"year":year,"records":len(records)}));db.commit()
    filename=f"Stock_FU_Mouvements_{year or 'All'}_{from_date or 'start'}_to_{to_date or 'end'}.csv"
    return StreamingResponse(iter(["\ufeff"+output.getvalue()]),media_type="text/csv; charset=utf-8",headers={"Content-Disposition":f'attachment; filename="{filename}"'})


def annual_rows(year:int,db:Session):
    cards=db.query(StockCard).filter(StockCard.is_active.is_(True)).order_by(StockCard.stock_card_number).all()
    result=[]
    start=date(year,1,1)
    for card in cards:
        opening=float(card.opening_quantity or 0)+float(db.query(func.coalesce(func.sum(StockMovement.quantity_in-StockMovement.quantity_out),0)).filter(StockMovement.stock_card_id==card.id,StockMovement.movement_date<start).scalar() or 0)
        row={"Stock Card":card.stock_card_number,"Donor":card.donor or "","Project Code":card.project_code or "","Item":card.item_name,"Specifications":card.specifications or card.item.specifications or "","Unit":card.unit,"Unit Price":card.unit_price,"Currency":card.currency,"Quantity minimum to have":card.minimum_quantity,"Quantities in Stock end of Y-1":opening,"months":[]}
        running=opening
        for month in range(1,13):
            total_in=float(db.query(func.coalesce(func.sum(StockMovement.quantity_in),0)).filter(StockMovement.stock_card_id==card.id,StockMovement.year==year,StockMovement.month_number==month).scalar() or 0)
            total_out=float(db.query(func.coalesce(func.sum(StockMovement.quantity_out),0)).filter(StockMovement.stock_card_id==card.id,StockMovement.year==year,StockMovement.month_number==month).scalar() or 0)
            running=running+total_in-total_out
            count=db.query(PhysicalInventoryCount).filter_by(stock_card_id=card.id,year=year,month_number=month).order_by(PhysicalInventoryCount.count_date.desc()).first()
            physical=count.physical_quantity if count else None
            difference=(physical-running) if physical is not None else None
            row["months"].append({"month":month,"name":month_name[month],"total_in":total_in,"total_out":total_out,"theoritical_quantity":running,"physical_inventory_quantity":physical,"difference":difference,"total_value":running*card.unit_price})
        result.append(row)
    return result


@router.get("/annual-summary/preview")
def preview_annual_summary(
    year:int=Query(...,ge=2000,le=2100),db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_annual_stock_summary")),
):
    rows=annual_rows(year,db);db.add(AuditLog(actor_id=actor.id,action="Annual Stock Summary previewed",details={"year":year,"records":len(rows)}));db.commit()
    return {"year":year,"items":rows}


def annual_headers():
    base=["Stock Card","Donor","Project Code","Item","Specifications","Unit","Unit Price","Currency","Quantity minimum to have","Quantities in Stock end of Y-1"]
    monthly=[]
    for month in range(1,13):
        name=month_name[month].upper()
        monthly += [f"{name} Total In",f"{name} Total Out",f"{name} Theoritical Quantity at end of the month",f"{name} Physical Inventory Quantity",f"{name} Difference",f"{name} Total Value End Of {month_name[month]}"]
    return base+monthly


def flatten_annual(row):
    values=[row[key] for key in annual_headers()[:10]]
    for month in row["months"]:
        values += [month["total_in"],month["total_out"],month["theoritical_quantity"],month["physical_inventory_quantity"],month["difference"],month["total_value"]]
    return values


@router.get("/annual-summary/export/csv")
def export_annual_csv(
    year:int=Query(...,ge=2000,le=2100),db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_annual_stock_summary")),
):
    rows=annual_rows(year,db);output=io.StringIO();writer=csv.writer(output);writer.writerow(annual_headers())
    for row in rows:writer.writerow(flatten_annual(row))
    db.add(AuditLog(actor_id=actor.id,action="Annual Stock Summary CSV exported",details={"year":year,"records":len(rows)}));db.commit()
    return StreamingResponse(iter(["\ufeff"+output.getvalue()]),media_type="text/csv; charset=utf-8",headers={"Content-Disposition":f'attachment; filename="Stock_FU_Annual_Summary_{year}.csv"'})


@router.get("/annual-summary/export/excel")
def export_annual_excel(
    year:int=Query(...,ge=2000,le=2100),db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_export_annual_stock_summary")),
):
    rows=annual_rows(year,db);workbook=Workbook();sheet=workbook.active;sheet.title="Summary"
    sheet.append(["STOCK FU : ANNUAL SUMMARY"]+[""]*(len(annual_headers())-1));sheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(annual_headers()))
    sheet["A1"].font=Font(bold=True,color="FFFFFF",size=16);sheet["A1"].fill=PatternFill("solid",fgColor="4B2828");sheet["A1"].alignment=Alignment(horizontal="center")
    sheet.append(annual_headers())
    for row in rows:sheet.append(flatten_annual(row))
    for cell in sheet[2]:cell.font=Font(bold=True);cell.fill=PatternFill("solid",fgColor="C69A3B");cell.alignment=Alignment(wrap_text=True,horizontal="center")
    sheet.freeze_panes="D3";sheet.auto_filter.ref=f"A2:{sheet.cell(2,len(annual_headers())).coordinate}"
    stream=io.BytesIO();workbook.save(stream);stream.seek(0)
    db.add(AuditLog(actor_id=actor.id,action="Annual Stock Summary Excel exported",details={"year":year,"records":len(rows)}));db.commit()
    return StreamingResponse(stream,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="Stock_FU_Annual_Summary_{year}.xlsx"'})
