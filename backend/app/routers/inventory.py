import csv
import io
import re
from datetime import date, datetime, timezone

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.inventory_template import INVENTORY_EXPORT_COLUMNS
from app.models import (
    AuditLog, InventoryItem, OrganizationSettings, SignatureEnvelope,
    SignatureRecipient, User,
)
from app.modules import permissions_for_role
from app.pagination import page_result
from app.schemas import (
    InventoryImportResult, InventoryImportRow, InventoryItemCreate,
    InventoryItemRead, InventoryItemUpdate, Page, UserRead,
)
from app.security import get_current_user, require_permission
from app.services.email_service import deliver_notification
from app.services.asset_form_pdf import build_asset_form_pdf
from app.services.sign_service import activate_recipient, audit, next_number, save_bytes, sha256_bytes
from app.stock_schemas import AssetFormPreview

router = APIRouter(
    prefix="/inventory",
    tags=["Inventory Management System"],
    dependencies=[Depends(require_permission("can_access_inventory"))],
)


def filtered_query(db, q=None, category=None, location=None, status=None, responsible_person=None, include_inactive=False):
    query = db.query(InventoryItem)
    if not include_inactive:
        query = query.filter(InventoryItem.is_active.is_(True))
    if q:
        term = f"%{q.strip()}%"
        query = query.filter(or_(
            InventoryItem.designation.ilike(term), InventoryItem.serial_number.ilike(term),
            InventoryItem.brand.ilike(term), InventoryItem.model.ilike(term),
            InventoryItem.number.ilike(term), InventoryItem.user_name.ilike(term),
        ))
    if category: query = query.filter(InventoryItem.category == category)
    if location: query = query.filter(InventoryItem.location == location)
    if status: query = query.filter(InventoryItem.status == status)
    if responsible_person: query = query.filter(InventoryItem.user_name.ilike(f"%{responsible_person}%"))
    return query


def apply_assignment_and_calculations(item: InventoryItem, db: Session):
    if item.assigned_user_id:
        assigned = db.get(User, item.assigned_user_id)
        if not assigned or not assigned.is_active:
            raise HTTPException(422, "Assigned user was not found")
        item.user_name = assigned.full_name
    if not item.purchasing_date:
        return
    parsed = None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(item.purchasing_date, pattern).date()
            break
        except ValueError:
            pass
    if not parsed:
        return
    today = date.today()
    months = max(0, (today.year - parsed.year) * 12 + today.month - parsed.month)
    item.months_since_purchasing = str(months)
    try:
        value = float(item.purchase_value_euros or 0)
        period = int(float(item.depreciation_period or 0))
        if period > 0:
            item.current_value_euros = f"{max(0, value * (1 - months / period)):.2f}"
    except (TypeError, ValueError):
        pass


@router.get("/assignees", response_model=list[UserRead])
def inventory_assignees(
    q: str | None = Query(None, max_length=120),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("can_export_asset_form")),
):
    query = db.query(User).filter(User.is_active.is_(True))
    if q and q.strip():
        term = f"%{q.strip()}%"
        query = query.filter(or_(
            User.full_name.ilike(term),
            User.email.ilike(term),
            User.department.ilike(term),
        ))
    return query.order_by(User.full_name).limit(limit).all()


@router.get("/my-assets", response_model=Page[InventoryItemRead])
def my_assets(
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db), user: User = Depends(get_current_user),
):
    query = db.query(InventoryItem).filter(
        InventoryItem.is_active.is_(True),
        InventoryItem.assigned_user_id == user.id,
    )
    return page_result(query.order_by(InventoryItem.updated_at.desc()), page, page_size)


def asset_form_data(user_id: int, db: Session):
    selected = db.get(User, user_id)
    if not selected or not selected.is_active:
        raise HTTPException(404, "User not found")
    assets = db.query(InventoryItem).filter(
        InventoryItem.is_active.is_(True),
        InventoryItem.assigned_user_id == selected.id,
    ).order_by(InventoryItem.designation).all()
    return selected, assets


def asset_form_reference(user_id: int, phase: str) -> str:
    return f"ASSET-FORM-{user_id}-{phase.upper()}"


def latest_asset_envelope(db: Session, user_id: int, phase: str):
    return db.query(SignatureEnvelope).filter(
        SignatureEnvelope.document_type == "asset_form",
        SignatureEnvelope.document_reference_id == asset_form_reference(user_id, phase),
    ).order_by(SignatureEnvelope.created_at.desc()).first()


def envelope_status(envelope: SignatureEnvelope | None) -> dict | None:
    if not envelope:
        return None
    return {
        "id": envelope.id,
        "envelope_id": envelope.envelope_id,
        "status": envelope.status,
        "completed_at": envelope.completed_at,
        "recipients": [{
            "user_id": item.user_id, "full_name": item.full_name,
            "role_name": item.role_name, "status": item.status,
            "signed_at": item.signed_at, "verification_number": item.verification_number,
        } for item in sorted(envelope.recipients, key=lambda row: row.routing_order)],
    }


def asset_signing_status(db: Session, user_id: int) -> dict:
    allocation = latest_asset_envelope(db, user_id, "allocation")
    returned = latest_asset_envelope(db, user_id, "return")
    if returned and returned.status == "completed":
        overall = "Received Back"
    elif returned and returned.status in {"pending", "in_progress"}:
        overall = "Return Signature Pending"
    elif allocation and allocation.status == "completed":
        overall = "Allocated"
    elif allocation and allocation.status in {"pending", "in_progress"}:
        overall = "Allocation Signature Pending"
    else:
        overall = "Not Signed"
    return {"overall_status": overall, "allocation": envelope_status(allocation), "return": envelope_status(returned)}


@router.get("/asset-form/preview", response_model=AssetFormPreview)
def preview_asset_form(
    user_id: int, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_export_asset_form")),
):
    selected, assets = asset_form_data(user_id, db)
    db.add(AuditLog(actor_id=actor.id, action="Asset form previewed", details={"user_id": user_id, "assets": len(assets)}))
    db.commit()
    preview = AssetFormPreview(
        user_id=selected.id, full_name=selected.full_name, email=selected.email,
        department=selected.department, generated_at=datetime.now(timezone.utc), assets=assets,
        signing=asset_signing_status(db, user_id),
    )
    return preview


@router.get("/asset-form/signing-status")
def asset_form_signing_status(
    user_id: int, db: Session = Depends(get_db),
    _: User = Depends(require_permission("can_export_asset_form")),
):
    selected, _ = asset_form_data(user_id, db)
    return {"user_id": selected.id, **asset_signing_status(db, user_id)}


@router.post("/asset-form/signing-request", status_code=201)
def create_asset_form_signing_request(
    user_id: int,
    background: BackgroundTasks,
    phase: str = Query(pattern="^(allocation|return)$"),
    db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_export_asset_form")),
):
    permissions = permissions_for_role(actor.role)
    if not {"can_create_signature_envelope", "can_send_signature_envelope"}.issubset(permissions):
        raise HTTPException(403, "You cannot create and send Asset Form signing requests")
    selected, assets = asset_form_data(user_id, db)
    if not assets:
        raise HTTPException(422, "No assigned assets were found for this staff member")
    if phase == "return":
        allocation = latest_asset_envelope(db, user_id, "allocation")
        if not allocation or allocation.status != "completed":
            raise HTTPException(409, "Allocation signatures must be completed before return signing")
    current = latest_asset_envelope(db, user_id, phase)
    if current and current.status in {"draft", "pending", "in_progress"}:
        raise HTTPException(409, f"{phase.title()} signing is already in progress")
    content = build_asset_form_pdf(selected, assets, None, asset_signing_status(db, user_id))
    stored = save_bytes("original-documents", content, ".pdf")
    document_hash = sha256_bytes(content)
    envelope = SignatureEnvelope(
        envelope_id=next_number(db, SignatureEnvelope, "envelope_id", "ENV"),
        document_type="asset_form",
        document_reference_id=asset_form_reference(user_id, phase),
        title=f"Asset Form - {selected.full_name} - {phase.title()}",
        subject=f"Asset Form {phase.title()} Signature Required",
        message=f"Please review and sign the {phase} section of your Asset Form.",
        status="draft", routing_mode="sequential",
        original_pdf_path=stored, original_pdf_hash=document_hash,
        current_document_hash=document_hash, created_by_id=actor.id,
    )
    db.add(envelope)
    db.flush()
    recipients = [(selected, f"Employee {phase}")]
    if actor.id != selected.id:
        logistics_role = "Logistics issued assets" if phase == "allocation" else "Logistics received assets"
        recipients.append((actor, logistics_role))
    for order, (recipient_user, role_name) in enumerate(recipients, 1):
        if phase == "allocation":
            signature_x = 0.47 if order == 1 else 0.68
        else:
            signature_x = 0.58 if order == 1 else 0.79
        db.add(SignatureRecipient(
            envelope_db_id=envelope.id, user_id=recipient_user.id,
            full_name=recipient_user.full_name, email=recipient_user.email,
            role_name=role_name, routing_order=order,
            signature_page=1, signature_x=signature_x, signature_y=0.46,
            signature_width=0.105, signature_height=0.12,
        ))
    db.flush()
    first = sorted(envelope.recipients, key=lambda item: item.routing_order)[0]
    signing_url, notification_id = activate_recipient(db, envelope, first)
    envelope.status = "pending"
    audit(db, envelope, "asset form envelope created and sent", user_id=actor.id, recipient_id=first.id, details={"phase": phase, "assets": len(assets)})
    db.add(AuditLog(actor_id=actor.id, action="Asset form signing requested", details={"user_id": user_id, "phase": phase, "envelope_id": envelope.envelope_id}))
    db.commit()
    if notification_id:
        background.add_task(deliver_notification, notification_id)
    return {
        "id": envelope.id, "envelope_id": envelope.envelope_id,
        "status": envelope.status, "phase": phase,
        "signing_url": signing_url if settings.environment == "development" else None,
    }


@router.get("/asset-form/export/pdf")
def export_asset_form(
    user_id: int, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_export_asset_form")),
):
    selected, assets = asset_form_data(user_id, db)
    organization = db.query(OrganizationSettings).first()
    content = build_asset_form_pdf(selected, assets, organization, asset_signing_status(db, user_id))
    db.add(AuditLog(actor_id=actor.id, action="Asset form PDF exported", details={"user_id": user_id, "assets": len(assets)}))
    db.commit()
    safe_name = re.sub(r"[^A-Za-z0-9_-]", "_", selected.full_name).strip("_")
    filename = f"Asset_Form_{safe_name}_{date.today():%Y-%m-%d}.pdf"
    return StreamingResponse(
        io.BytesIO(content), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"', "X-Content-Type-Options": "nosniff"},
    )


@router.get("/items/export/csv")
def export_inventory(
    q: str | None = None, category: str | None = None, location: str | None = None,
    status: str | None = None, responsible_person: str | None = None,
    db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_export_inventory")),
):
    items = filtered_query(db, q, category, location, status, responsible_person).order_by(InventoryItem.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([header for header, _ in INVENTORY_EXPORT_COLUMNS])
    for item in items:
        writer.writerow([getattr(item, field) or "" for _, field in INVENTORY_EXPORT_COLUMNS])
    db.add(AuditLog(actor_id=actor.id, action="Inventory CSV exported", details={"records": len(items)}))
    db.commit()
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]), media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="inventory-register.csv"'},
    )


def _normalized_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


INVENTORY_IMPORT_FIELDS = {
    _normalized_header(header): field for header, field in INVENTORY_EXPORT_COLUMNS
}


def _parse_inventory_upload(content: bytes, filename: str) -> list[dict]:
    extension = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if extension == ".csv":
        try:
            return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
        except UnicodeDecodeError:
            raise HTTPException(422, "CSV files must use UTF-8 encoding")
    if extension == ".xlsx":
        if not content.startswith(b"PK"):
            raise HTTPException(415, "File contents do not match XLSX format")
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in rows[1:]
            ]
        except Exception as exc:
            raise HTTPException(422, f"Could not read XLSX file: {str(exc)[:120]}")
    raise HTTPException(415, "Upload a CSV or XLSX inventory file")


def _clean_import_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text_value = str(value).strip()
    return text_value or None


def _row_payload(raw: dict) -> dict:
    normalized = {_normalized_header(key): _clean_import_value(value) for key, value in raw.items()}
    payload = {
        field: normalized.get(header_key)
        for header_key, field in INVENTORY_IMPORT_FIELDS.items()
        if field != "logistics_code" and normalized.get(header_key) is not None
    }
    logistics_code = normalized.get("logistics_code")
    if logistics_code:
        parts = [part.strip() for part in re.split(r"[/.]", logistics_code) if part.strip()]
        if len(parts) != 5:
            raise ValueError("Logistics Code must contain Country/Project/Category/Sub-Category/Number")
        for field, value in zip(("country", "project", "category", "sub_category", "number"), parts):
            payload.setdefault(field, value)
    return payload


def _duplicate_inventory_item(db: Session, payload: InventoryItemCreate) -> str | None:
    if payload.serial_number:
        serial_exists = db.query(InventoryItem.id).filter(
            func.lower(InventoryItem.serial_number) == payload.serial_number.lower()
        ).first()
        if serial_exists:
            return f"Serial number already exists: {payload.serial_number}"
    code_parts = [payload.country, payload.project, payload.category, payload.sub_category, payload.number]
    if all(code_parts):
        code_exists = db.query(InventoryItem.id).filter(
            func.lower(InventoryItem.country) == payload.country.lower(),
            func.lower(InventoryItem.project) == payload.project.lower(),
            func.lower(InventoryItem.category) == payload.category.lower(),
            func.lower(InventoryItem.sub_category) == payload.sub_category.lower(),
            func.lower(InventoryItem.number) == payload.number.lower(),
        ).first()
        if code_exists:
            return f"Logistics code already exists: {'/'.join(code_parts)}"
    return None


@router.post("/items/import", response_model=InventoryImportResult)
async def import_inventory(
    file: UploadFile = File(...),
    confirm: bool = False,
    db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_import_inventory")),
):
    content = await file.read(5_242_881)
    await file.close()
    if len(content) > 5_242_880:
        raise HTTPException(413, "Inventory import files must be smaller than 5 MB")
    parsed = _parse_inventory_upload(content, file.filename or "")
    if len(parsed) > 2_000:
        raise HTTPException(422, "Inventory import is limited to 2,000 rows per file")

    rows: list[InventoryImportRow] = []
    valid = invalid = imported = skipped = 0
    seen_serials: set[str] = set()
    seen_codes: set[str] = set()
    for row_number, raw in enumerate(parsed, 2):
        warnings: list[str] = []
        try:
            data = _row_payload(raw)
            user_value = str(data.pop("user_name", "") or "").strip()
            assigned = None
            if user_value:
                assigned = db.query(User).filter(
                    User.is_active.is_(True),
                    or_(
                        func.lower(User.email) == user_value.lower(),
                        func.lower(User.full_name) == user_value.lower(),
                    ),
                ).first()
                if assigned:
                    data["assigned_user_id"] = assigned.id
                    data["user_name"] = assigned.full_name
                else:
                    data["user_name"] = user_value
                    warnings.append(f'User "{user_value}" was not matched; item will remain unassigned')
            payload = InventoryItemCreate.model_validate(data)
            serial_key = (payload.serial_number or "").lower()
            code_parts = [payload.country, payload.project, payload.category, payload.sub_category, payload.number]
            code_key = "/".join(code_parts).lower() if all(code_parts) else ""
            if serial_key and serial_key in seen_serials:
                raise ValueError(f"Duplicate serial number in upload: {payload.serial_number}")
            if code_key and code_key in seen_codes:
                raise ValueError(f"Duplicate logistics code in upload: {code_key.upper()}")
            duplicate = _duplicate_inventory_item(db, payload)
            if duplicate:
                skipped += 1
                rows.append(InventoryImportRow(
                    row=row_number, status="skipped", detail=duplicate,
                    warnings=warnings, data=payload.model_dump(mode="json"),
                ))
                continue
            if serial_key:
                seen_serials.add(serial_key)
            if code_key:
                seen_codes.add(code_key)
            valid += 1
            if confirm:
                item = InventoryItem(**payload.model_dump(), created_by_id=actor.id)
                apply_assignment_and_calculations(item, db)
                db.add(item)
                imported += 1
            rows.append(InventoryImportRow(
                row=row_number, status="imported" if confirm else "valid",
                warnings=warnings,
                data={
                    **payload.model_dump(mode="json"),
                    "logistics_code": code_key.upper() if code_key else "",
                },
            ))
        except (ValidationError, ValueError) as exc:
            invalid += 1
            rows.append(InventoryImportRow(
                row=row_number, status="invalid", detail=str(exc)[:500], data=dict(raw),
            ))
    if confirm:
        db.add(AuditLog(
            actor_id=actor.id,
            action="Inventory CSV/Excel imported",
            details={
                "filename": file.filename, "total": len(parsed), "imported": imported,
                "skipped": skipped, "errors": invalid,
            },
        ))
        db.commit()
    return InventoryImportResult(
        total=len(parsed), valid=valid, invalid=invalid,
        imported=imported, skipped=skipped, rows=rows,
    )


@router.get("/items", response_model=Page[InventoryItemRead])
def list_items(
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100),
    q: str | None = Query(None, max_length=200), category: str | None = None,
    location: str | None = None, status: str | None = None,
    responsible_person: str | None = None, include_inactive: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("can_manage_inventory")),
):
    query = filtered_query(db, q, category, location, status, responsible_person, include_inactive)
    return page_result(query.order_by(InventoryItem.updated_at.desc()), page, page_size)


@router.post("/items", response_model=InventoryItemRead, status_code=201)
def create_item(
    data: InventoryItemCreate, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_inventory")),
):
    item = InventoryItem(**data.model_dump(), created_by_id=actor.id)
    apply_assignment_and_calculations(item, db)
    db.add(item); db.flush()
    db.add(AuditLog(actor_id=actor.id, action="Inventory item created", details={"inventory_item_id": item.id}))
    db.commit(); db.refresh(item)
    return item


@router.get("/items/{item_id}", response_model=InventoryItemRead)
def get_item(
    item_id: int, db: Session = Depends(get_db),
    _: User = Depends(require_permission("can_manage_inventory")),
):
    item = db.get(InventoryItem, item_id)
    if not item: raise HTTPException(404, "Inventory item not found")
    return item


@router.put("/items/{item_id}", response_model=InventoryItemRead)
def update_item(
    item_id: int, data: InventoryItemUpdate, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_inventory")),
):
    item = db.get(InventoryItem, item_id)
    if not item: raise HTTPException(404, "Inventory item not found")
    changes = data.model_dump(exclude_unset=True)
    for field, value in changes.items(): setattr(item, field, value)
    apply_assignment_and_calculations(item, db)
    db.add(AuditLog(actor_id=actor.id, action="Inventory item updated", details={"inventory_item_id": item.id, "fields": sorted(changes)}))
    db.commit(); db.refresh(item)
    return item


@router.patch("/items/{item_id}/deactivate", response_model=InventoryItemRead)
def deactivate_item(
    item_id: int, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_inventory")),
):
    item = db.get(InventoryItem, item_id)
    if not item: raise HTTPException(404, "Inventory item not found")
    item.is_active = False
    db.add(AuditLog(actor_id=actor.id, action="Inventory item deactivated", details={"inventory_item_id": item.id}))
    db.commit(); db.refresh(item)
    return item
