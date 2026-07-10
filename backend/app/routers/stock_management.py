import csv
import io
import re
from datetime import date

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    AuditLog, PhysicalInventoryCount, StockCard, StockCategory, StockItem,
    StockMovement, User,
)
from app.pagination import page_result
from app.schemas import Page, StockItemCreate, StockItemRead
from app.security import require_permission
from app.stock_schemas import (
    ImportPreview, ImportRow, PhysicalCountCreate, StockCardCreate, StockCardRead,
    StockCardUpdate, StockCategoryCreate, StockCategoryRead, StockCategoryUpdate,
    StockMovementCreate, StockMovementRead,
)

router = APIRouter(
    prefix="/stock",
    tags=["Stock Management"],
    dependencies=[Depends(require_permission("can_access_stock"))],
)


def sync_item_status(item: StockItem):
    if not item.is_active: item.status = "Inactive"
    elif item.quantity_available <= 0: item.status = "Out of Stock"
    elif item.quantity_available <= item.low_stock_threshold: item.status = "Low Stock"
    else: item.status = "Available"


@router.get("/categories", response_model=list[StockCategoryRead])
def list_categories(
    include_inactive: bool = False, db: Session = Depends(get_db),
):
    query = db.query(StockCategory)
    if not include_inactive:
        query = query.filter(StockCategory.is_active.is_(True))
    return query.order_by(StockCategory.display_order, StockCategory.name).all()


@router.post("/categories", response_model=StockCategoryRead, status_code=201)
def create_category(
    data: StockCategoryCreate, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_stock_categories")),
):
    if db.query(StockCategory).filter(StockCategory.name.ilike(data.name.strip())).first():
        raise HTTPException(409, "Stock category already exists")
    category = StockCategory(**data.model_dump(), created_by_id=actor.id, updated_by_id=actor.id)
    db.add(category); db.flush()
    db.add(AuditLog(actor_id=actor.id, action="Stock category created", details={"stock_category_id": category.id}))
    db.commit(); db.refresh(category)
    return category


@router.get("/categories/{category_id}", response_model=StockCategoryRead)
def get_category(category_id: int, db: Session = Depends(get_db)):
    category = db.get(StockCategory, category_id)
    if not category: raise HTTPException(404, "Stock category not found")
    return category


@router.put("/categories/{category_id}", response_model=StockCategoryRead)
def update_category(
    category_id: int, data: StockCategoryUpdate, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_stock_categories")),
):
    category = db.get(StockCategory, category_id)
    if not category: raise HTTPException(404, "Stock category not found")
    changes = data.model_dump(exclude_unset=True)
    for field, value in changes.items(): setattr(category, field, value)
    category.updated_by_id = actor.id
    if "name" in changes:
        for item in category.items: item.category = category.name
    db.add(AuditLog(actor_id=actor.id, action="Stock category updated", details={"stock_category_id": category.id, "fields": sorted(changes)}))
    db.commit(); db.refresh(category)
    return category


@router.patch("/categories/{category_id}/deactivate", response_model=StockCategoryRead)
def deactivate_category(
    category_id: int, db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_manage_stock_categories")),
):
    category = db.get(StockCategory, category_id)
    if not category: raise HTTPException(404, "Stock category not found")
    category.is_active = False; category.updated_by_id = actor.id
    db.add(AuditLog(actor_id=actor.id, action="Stock category deactivated", details={"stock_category_id": category.id}))
    db.commit(); db.refresh(category)
    return category


@router.get("/categories/{category_id}/items", response_model=Page[StockItemRead])
def category_items(
    category_id: int, page: int = Query(1, ge=1), page_size: int = Query(24, ge=1, le=100),
    q: str | None = None, status: str | None = None, db: Session = Depends(get_db),
):
    category = db.get(StockCategory, category_id)
    if not category or not category.is_active: raise HTTPException(404, "Stock category not found")
    query = db.query(StockItem).filter(StockItem.category_id == category_id, StockItem.is_active.is_(True))
    if q:
        term = f"%{q.strip()}%"
        query = query.filter(or_(StockItem.item_name.ilike(term), StockItem.specifications.ilike(term)))
    if status: query = query.filter(StockItem.status == status)
    return page_result(query.order_by(StockItem.item_name), page, page_size)


def parse_upload(content: bytes, filename: str) -> list[dict]:
    extension = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if extension == ".csv":
        try:
            return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
        except UnicodeDecodeError:
            raise HTTPException(422, "CSV must use UTF-8 encoding")
    if extension == ".xlsx":
        if not content.startswith(b"PK"):
            raise HTTPException(415, "File contents do not match XLSX format")
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows: return []
            headers = [str(value or "").strip() for value in rows[0]]
            return [{headers[i]: value for i, value in enumerate(row) if i < len(headers)} for row in rows[1:]]
        except Exception as exc:
            raise HTTPException(422, f"Could not read XLSX file: {str(exc)[:120]}")
    raise HTTPException(415, "Upload a CSV or XLSX file")


def normalized(row: dict) -> dict:
    return {re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower()).strip("_"): value for key, value in row.items()}


@router.post("/items/import", response_model=ImportPreview)
async def import_items(
    file: UploadFile = File(...), confirm: bool = False,
    db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_import_stock")),
):
    content = await file.read(5_242_881); await file.close()
    if len(content) > 5_242_880: raise HTTPException(413, "Import files must be smaller than 5 MB")
    parsed = parse_upload(content, file.filename or "")
    results=[]; valid=invalid=created=0
    for index, raw in enumerate(parsed[:2000], 2):
        row=normalized(raw)
        data={
            "item_name": row.get("item_name") or row.get("item"),
            "category": row.get("category") or row.get("item_type") or "Other",
            "specifications": row.get("specifications") or "",
            "quantity_available": row.get("quantity_available") or row.get("quantity") or 0,
            "unit": row.get("unit") or "piece",
            "location": row.get("location") or row.get("base") or "Coordination",
            "unit_price": row.get("unit_price") or 0,
            "currency": row.get("currency") or "AFN",
            "donor": row.get("donor") or None,
            "project_code": row.get("project_code") or None,
            "notes": row.get("comments") or row.get("notes") or None,
        }
        try:
            payload=StockItemCreate.model_validate(data)
            valid+=1
            if confirm:
                category=db.query(StockCategory).filter(StockCategory.name.ilike(payload.category)).first()
                if not category:
                    category=StockCategory(name=payload.category,description="Imported category",display_order=999,created_by_id=actor.id,updated_by_id=actor.id)
                    db.add(category);db.flush()
                item=StockItem(**payload.model_dump(exclude={"category_id"}),category_id=category.id)
                sync_item_status(item);db.add(item);created+=1
            results.append(ImportRow(row=index,status="valid" if not confirm else "created",data=payload.model_dump(mode="json")))
        except Exception as exc:
            invalid+=1;results.append(ImportRow(row=index,status="invalid",detail=str(exc)[:300],data=data))
    if confirm:
        db.add(AuditLog(actor_id=actor.id,action="Stock item imported",details={"created":created,"invalid":invalid}))
        db.commit()
    return ImportPreview(valid=valid,invalid=invalid,created=created,rows=results)


def card_number(base: str, location: str, number: str) -> str:
    base=base.upper();location=location.upper()
    if not re.fullmatch(r"[A-Z]{3}",base): raise HTTPException(422,"Base must contain exactly 3 letters")
    if not re.fullmatch(r"[A-Z]{2}[0-9]",location): raise HTTPException(422,"Location must contain exactly 2 letters and 1 number")
    if not re.fullmatch(r"[0-9]{3}",number): raise HTTPException(422,"Number must contain exactly 3 digits")
    return f"{base}.{location}.{number}"


@router.get("/stock-cards/generate-number")
def generate_card_number(
    base: str, location: str, db: Session = Depends(get_db),
    _: User = Depends(require_permission("can_manage_stock_cards")),
):
    base=base.upper();location=location.upper();card_number(base,location,"001")
    cards=db.query(StockCard).filter(StockCard.base==base,StockCard.storage_location==location).all()
    next_number=max([int(card.sequence_number) for card in cards] or [0])+1
    if next_number>999: raise HTTPException(409,"No stock card numbers remain for this base and location")
    number=f"{next_number:03d}"
    return {"base":base,"storage_location":location,"sequence_number":number,"stock_card_number":card_number(base,location,number)}


@router.get("/stock-cards/export/csv")
def export_cards(
    db: Session = Depends(get_db),
    actor: User = Depends(require_permission("can_export_stock_cards")),
):
    cards=db.query(StockCard).order_by(StockCard.stock_card_number).all()
    output=io.StringIO();writer=csv.writer(output)
    writer.writerow(["Base","Location of Storage","Number","Stock Card Number","Donor","Project Code","Item","Specifications","Unit","Expiration Date","Unit Price","Currency","Comments"])
    for card in cards:
        writer.writerow([card.base,card.storage_location,card.sequence_number,card.stock_card_number,card.donor or "",card.project_code or "",card.item_name,card.specifications or "",card.unit,card.expiration_date.isoformat() if card.expiration_date else "",card.unit_price,card.currency,card.comments or ""])
    db.add(AuditLog(actor_id=actor.id,action="Stock card CSV exported",details={"records":len(cards)}));db.commit()
    return StreamingResponse(iter(["\ufeff"+output.getvalue()]),media_type="text/csv; charset=utf-8",headers={"Content-Disposition":'attachment; filename="Stock_Card_Register.csv"'})


@router.get("/stock-cards", response_model=Page[StockCardRead])
def list_cards(
    page:int=Query(1,ge=1),page_size:int=Query(25,ge=1,le=100),q:str|None=None,
    db:Session=Depends(get_db),_:User=Depends(require_permission("can_access_stock_cards")),
):
    query=db.query(StockCard)
    if q:
        term=f"%{q}%";query=query.join(StockItem).filter(or_(StockCard.stock_card_number.ilike(term),StockItem.item_name.ilike(term)))
    return page_result(query.order_by(StockCard.stock_card_number),page,page_size)


@router.post("/stock-cards", response_model=StockCardRead, status_code=201)
def create_card(
    data:StockCardCreate,db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_manage_stock_cards")),
):
    number=card_number(data.base,data.storage_location,data.sequence_number)
    if db.query(StockCard).filter_by(stock_card_number=number).first(): raise HTTPException(409,"Stock card number already exists")
    item=db.get(StockItem,data.stock_item_id)
    if not item: raise HTTPException(404,"Stock item not found")
    values=data.model_dump();values.update(base=data.base.upper(),storage_location=data.storage_location.upper(),stock_card_number=number,created_by_id=actor.id)
    card=StockCard(**values);db.add(card);db.flush()
    db.add(AuditLog(actor_id=actor.id,action="Stock card created",details={"stock_card_id":card.id,"stock_card_number":number}))
    db.commit();db.refresh(card);return card


@router.get("/stock-cards/{card_id}", response_model=StockCardRead)
def get_card(card_id:int,db:Session=Depends(get_db),_:User=Depends(require_permission("can_access_stock_cards"))):
    card=db.get(StockCard,card_id)
    if not card:raise HTTPException(404,"Stock card not found")
    return card


@router.put("/stock-cards/{card_id}", response_model=StockCardRead)
def update_card(card_id:int,data:StockCardUpdate,db:Session=Depends(get_db),actor:User=Depends(require_permission("can_manage_stock_cards"))):
    card=db.get(StockCard,card_id)
    if not card:raise HTTPException(404,"Stock card not found")
    changes=data.model_dump(exclude_unset=True)
    for field,value in changes.items():setattr(card,field,value)
    db.add(AuditLog(actor_id=actor.id,action="Stock card updated",details={"stock_card_id":card.id,"fields":sorted(changes)}));db.commit();db.refresh(card);return card


@router.patch("/stock-cards/{card_id}/deactivate", response_model=StockCardRead)
def deactivate_card(card_id:int,db:Session=Depends(get_db),actor:User=Depends(require_permission("can_manage_stock_cards"))):
    card=db.get(StockCard,card_id)
    if not card:raise HTTPException(404,"Stock card not found")
    card.is_active=False;db.add(AuditLog(actor_id=actor.id,action="Stock card deactivated",details={"stock_card_id":card.id}));db.commit();db.refresh(card);return card


@router.get("/movements", response_model=Page[StockMovementRead])
def list_movements(
    page:int=Query(1,ge=1),page_size:int=Query(25,ge=1,le=100),stock_item_id:int|None=None,
    year:int|None=None,month_number:int|None=None,db:Session=Depends(get_db),
    _:User=Depends(require_permission("can_access_stock_cards")),
):
    query=db.query(StockMovement)
    if stock_item_id:query=query.filter(StockMovement.stock_item_id==stock_item_id)
    if year:query=query.filter(StockMovement.year==year)
    if month_number:query=query.filter(StockMovement.month_number==month_number)
    return page_result(query.order_by(StockMovement.movement_date.desc(),StockMovement.id.desc()),page,page_size)


@router.post("/movements", response_model=StockMovementRead, status_code=201)
def create_movement(
    data:StockMovementCreate,db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_manage_stock")),
):
    item=db.get(StockItem,data.stock_item_id)
    if not item:raise HTTPException(404,"Stock item not found")
    if data.quantity_in<=0 and data.quantity_out<=0:raise HTTPException(422,"Enter an IN or OUT quantity")
    net=data.quantity_in-data.quantity_out
    if item.quantity_available+net<0:raise HTTPException(409,"Movement would make stock negative")
    item.quantity_available+=int(net);sync_item_status(item)
    movement=StockMovement(**data.model_dump(),quantity_change=int(net),month_number=data.movement_date.month,year=data.movement_date.year,performed_by_id=actor.id)
    db.add(movement);db.flush()
    db.add(AuditLog(actor_id=actor.id,action="Stock quantity changed",details={"stock_movement_id":movement.id,"stock_item_id":item.id,"change":net}))
    db.commit();db.refresh(movement);return movement


@router.post("/physical-counts", status_code=201)
def create_physical_count(
    data:PhysicalCountCreate,db:Session=Depends(get_db),
    actor:User=Depends(require_permission("can_manage_stock")),
):
    if not db.get(StockItem,data.stock_item_id):raise HTTPException(404,"Stock item not found")
    count=PhysicalInventoryCount(**data.model_dump(),year=data.count_date.year,month_number=data.count_date.month,counted_by_id=actor.id)
    db.add(count);db.commit();db.refresh(count)
    return {"id":count.id,"year":count.year,"month_number":count.month_number,"physical_quantity":count.physical_quantity}
